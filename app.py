from flask import Flask, render_template, request, jsonify, send_file
from database import db, User, Attendance
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from werkzeug.utils import secure_filename
import os

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///attendance.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = 'static/photos'
app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024  # 5MB max file size

# Ensure upload folder exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

db.init_app(app)

# Committee list
COMMITTEES = [
    'Executive Board',
    'Externals',
    'Internals',
    'Visuals and Marketing',
    'Finance',
    'Admin & Productions'
]

with app.app_context():
    db.create_all()
    # Add sample users if database is empty
    if User.query.count() == 0:
        sample_users = [
            User(id_number='20212345', full_name='Juan Dela Cruz', birthday='01-15', committee='Executive Board'),
            User(id_number='20212346', full_name='Maria Santos', birthday='03-22', committee='Externals'),
            User(id_number='20212347', full_name='Pedro Reyes', birthday='07-08', committee='Internals')
        ]
        for user in sample_users:
            db.session.add(user)
        db.session.commit()
        print("Sample users added to database.")

@app.route('/')
def index():
    return render_template('dashboard.html', committees=COMMITTEES)

# ============ ATTENDANCE ENDPOINTS ============

@app.route('/api/attendance', methods=['POST'])
def toggle_attendance():
    """Single endpoint to handle both time in and time out"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
            
        id_number = data.get('id_number', '').strip()
        
        if not id_number:
            return jsonify({'error': 'ID Number is required'}), 400
        
        user = User.query.filter_by(id_number=id_number).first()
        if not user:
            return jsonify({'error': 'ID Number not found in the system'}), 404
        
        # Check latest attendance record
        latest = Attendance.query.filter_by(user_id=user.id).order_by(Attendance.timestamp.desc()).first()
        
        # Determine action: if last action was 'In' or no record, do 'Out', otherwise do 'In'
        if latest and latest.event_type == 'In':
            # Time Out
            attendance = Attendance(user_id=user.id, event_type='Out')
            db.session.add(attendance)
            db.session.commit()
            
            return jsonify({
                'success': True,
                'action': 'out',
                'message': f"Goodbye, {user.full_name}! You are timed out.",
                'timestamp': attendance.timestamp.strftime('%Y-%m-%d %I:%M:%S %p'),
                'user': user.to_dict()
            })
        else:
            # Time In
            attendance = Attendance(user_id=user.id, event_type='In')
            db.session.add(attendance)
            db.session.commit()
            
            # Check for birthday
            today = datetime.now()
            current_month_day = today.strftime('%m-%d')
            is_birthday = user.birthday == current_month_day
            
            message = f"Happy Birthday, {user.full_name}! You are timed in." if is_birthday else f"Welcome, {user.full_name}! You are timed in."
            
            return jsonify({
                'success': True,
                'action': 'in',
                'message': message,
                'is_birthday': is_birthday,
                'timestamp': attendance.timestamp.strftime('%Y-%m-%d %I:%M:%S %p'),
                'user': user.to_dict()
            })
            
    except Exception as e:
        print(f"Error in toggle_attendance: {str(e)}")
        db.session.rollback()
        return jsonify({'error': f'Server error: {str(e)}'}), 500

@app.route('/api/active-users')
def get_active_users():
    """Get all currently logged in users grouped by committee"""
    try:
        active_users = {}
        
        for committee in COMMITTEES:
            users_in_committee = User.query.filter_by(committee=committee).all()
            active_in_committee = []
            
            for user in users_in_committee:
                latest = Attendance.query.filter_by(user_id=user.id).order_by(Attendance.timestamp.desc()).first()
                if latest and latest.event_type == 'In':
                    active_in_committee.append({
                        'id': user.id,
                        'id_number': user.id_number,
                        'full_name': user.full_name,
                        'photo_filename': user.photo_filename,
                        'time_in': latest.timestamp.strftime('%I:%M %p')
                    })
            
            active_users[committee] = active_in_committee
        
        return jsonify(active_users)
    except Exception as e:
        print(f"Error in get_active_users: {str(e)}")
        return jsonify({'error': str(e)}), 500

# ============ USER MANAGEMENT ENDPOINTS ============

@app.route('/api/users', methods=['GET'])
def get_users():
    """Get all users"""
    try:
        users = User.query.all()
        return jsonify([user.to_dict() for user in users])
    except Exception as e:
        print(f"Error in get_users: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/users/search', methods=['GET'])
def search_users():
    """Search users by name or ID"""
    try:
        query = request.args.get('q', '').strip()
        if not query:
            return jsonify([])
        
        users = User.query.filter(
            (User.full_name.ilike(f'%{query}%')) | 
            (User.id_number.ilike(f'%{query}%'))
        ).all()
        
        return jsonify([user.to_dict() for user in users])
    except Exception as e:
        print(f"Error in search_users: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/users', methods=['POST'])
def add_user():
    """Add a new user"""
    try:
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['id_number', 'full_name', 'birthday', 'committee']
        for field in required_fields:
            if field not in data or not data[field]:
                return jsonify({'error': f'{field} is required'}), 400
        
        # Check if user already exists
        existing = User.query.filter_by(id_number=data['id_number']).first()
        if existing:
            return jsonify({'error': 'User with this ID number already exists'}), 400
        
        # Create new user
        user = User(
            id_number=data['id_number'],
            full_name=data['full_name'],
            birthday=data['birthday'],
            committee=data['committee'],
            photo_filename=data.get('photo_filename')
        )
        
        db.session.add(user)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'User {user.full_name} added successfully',
            'user': user.to_dict()
        })
        
    except Exception as e:
        print(f"Error in add_user: {str(e)}")
        db.session.rollback()
        return jsonify({'error': f'Server error: {str(e)}'}), 500

@app.route('/api/users/<int:user_id>', methods=['DELETE'])
def delete_user(user_id):
    """Delete a user"""
    try:
        user = User.query.get(user_id)
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        # Delete associated photo if exists
        if user.photo_filename:
            photo_path = os.path.join(app.config['UPLOAD_FOLDER'], user.photo_filename)
            if os.path.exists(photo_path):
                os.remove(photo_path)
        
        db.session.delete(user)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'User {user.full_name} deleted successfully'
        })
        
    except Exception as e:
        print(f"Error in delete_user: {str(e)}")
        db.session.rollback()
        return jsonify({'error': f'Server error: {str(e)}'}), 500

@app.route('/api/users/<int:user_id>/photo', methods=['POST'])
def upload_photo(user_id):
    """Upload user photo"""
    try:
        user = User.query.get(user_id)
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        if 'photo' not in request.files:
            return jsonify({'error': 'No photo file provided'}), 400
        
        file = request.files['photo']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if file:
            # Delete old photo if exists
            if user.photo_filename:
                old_photo = os.path.join(app.config['UPLOAD_FOLDER'], user.photo_filename)
                if os.path.exists(old_photo):
                    os.remove(old_photo)
            
            # Save new photo
            filename = secure_filename(f"{user.id_number}_{file.filename}")
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            
            user.photo_filename = filename
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': 'Photo uploaded successfully',
                'photo_filename': filename
            })
            
    except Exception as e:
        print(f"Error in upload_photo: {str(e)}")
        db.session.rollback()
        return jsonify({'error': f'Server error: {str(e)}'}), 500

# ============ EXPORT ENDPOINT ============

@app.route('/api/export-dtr')
def export_dtr():
    """Export attendance logs as Excel DTR format"""
    try:
        attendance_records = db.session.query(
            User.full_name,
            User.committee,
            Attendance.timestamp,
            Attendance.event_type
        ).join(User).order_by(Attendance.timestamp.desc()).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "DTR Report"
        
        header_fill = PatternFill(start_color="00693C", end_color="00693C", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        headers = ['Date', 'Full Name', 'Committee', 'Time In', 'Time Out', 'Total Hours']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        dtr_data = {}
        for name, committee, timestamp, event_type in attendance_records:
            date_key = timestamp.strftime('%Y-%m-%d')
            user_date_key = f"{name}_{date_key}"
            
            if user_date_key not in dtr_data:
                dtr_data[user_date_key] = {
                    'date': date_key,
                    'name': name,
                    'committee': committee,
                    'time_in': None,
                    'time_out': None
                }
            
            if event_type == 'In':
                if dtr_data[user_date_key]['time_in'] is None:
                    dtr_data[user_date_key]['time_in'] = timestamp
            else:
                if dtr_data[user_date_key]['time_out'] is None:
                    dtr_data[user_date_key]['time_out'] = timestamp
        
        row = 2
        for key in sorted(dtr_data.keys(), reverse=True):
            record = dtr_data[key]
            
            ws.cell(row=row, column=1, value=record['date'])
            ws.cell(row=row, column=2, value=record['name'])
            ws.cell(row=row, column=3, value=record['committee'])
            ws.cell(row=row, column=4, value=record['time_in'].strftime('%I:%M %p') if record['time_in'] else '')
            ws.cell(row=row, column=5, value=record['time_out'].strftime('%I:%M %p') if record['time_out'] else '')
            
            if record['time_in'] and record['time_out']:
                delta = record['time_out'] - record['time_in']
                hours = delta.total_seconds() / 3600
                ws.cell(row=row, column=6, value=f"{hours:.2f}")
            else:
                ws.cell(row=row, column=6, value='')
            
            row += 1
        
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        
        filename = f"DTR_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        exports_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'exports')
        os.makedirs(exports_dir, exist_ok=True)
        
        filepath = os.path.join(exports_dir, filename)
        wb.save(filepath)
        
        return send_file(
            filepath,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        print(f"Error in export_dtr: {str(e)}")
        return jsonify({'error': f'Export failed: {str(e)}'}), 500

if __name__ == '__main__':
    app.run(debug=True, host='127.0.0.1', port=5000)
